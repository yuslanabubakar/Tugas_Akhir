from openpyxl import load_workbook
import function as fc
import time
import numpy as np

start_time = time.time()
#matriksTarget = [[1,0,0,0,0,0,0],
#                 [0,1,0,0,0,0,0],
#                 [0,0,1,0,0,0,0],
#                 [0,0,0,1,0,0,0],
#                 [0,0,0,0,1,0,0],
#                 [0,0,0,0,0,1,0],
#                 [0,0,0,0,0,0,1]]
matriksTarget = [[0,0,1],
                 [0,1,0],
                 [0,1,1],
                 [1,0,0],
                 [1,0,1],
                 [1,1,0],
                 [1,1,1]]
#np.savetxt('matriksTarget.txt',matriksTarget)
print 'Open File...'
wb = load_workbook('dataMulti.xlsx')
sheet = wb.active
dataSet = fc.getData(sheet)

hammingLoss = []
num_folds = 5
subset_size = len(dataSet) / num_folds
for i in range(num_folds): #K-Cross Validation
    print 'K-Cross = ',(i+1)
    testing = dataSet[i*subset_size:][:subset_size]
    training = dataSet[:i*subset_size]+dataSet[(i+1)*subset_size:]
    dataPre = fc.preprocessing(training)
    print 'Counting document for each class...'
    docEachClass = fc.docEachClass(matriksTarget,training,dataPre)
    print 'Counting document for each class given word...'
    Fw = fc.docGivenWord(matriksTarget,dataPre,training)
    print 'Counting Probability for Each Class Given Word w...'
    piW = fc.probEachClassGivenWord(Fw,docEachClass,matriksTarget)
    print 'Counting Information Gain for Each Class and Word...'
    igW = fc.informationGain(docEachClass,training,piW,dataPre)
    
    #get words if igW value > threshold
    igWThreshold = [x[0] for x in igW if x[1] > 0.60]
#    igWThreshold = []
#    for i in ig:
#        igWThreshold.append(i[0])
    #np.savetxt("igWThreshold.txt", igWThreshold, delimiter=",", fmt="%s")
    print 'Counting TF...'
    tf = fc.tf(igWThreshold,training)
    print 'Counting IDF...'
    idf = fc.idf(tf,training)
    print 'Counting TFxIDF...'
    tfIDF = fc.tfIDF(tf,idf,training)
    
    #start training ann (artificial neural network)
    input_p = len(tfIDF[0][1])
    hidden_p = 10
    output_p = 3
    lr = 0.1
    epoch = 1000
    mseStandar = 0.01
    #i_param = [input_p,hidden_p,output_p,lr,epoch,mseStandar]
    #np.savetxt('inputParameters.txt',i_param)
    W1,W2,B1,B2 = fc.training(input_p,hidden_p,output_p,lr,epoch,mseStandar,tfIDF,training)
    
    #start testing
    print 'Now is testing the data...'
#    wbTest = load_workbook('dataTest.xlsx')
#    sheetTest = wbTest.active
#    dataSetTest = fc.getData(sheetTest)
    print 'Counting TF...'
    tfTest = fc.tf(igWThreshold,testing)
    print 'Counting IDF...'
    idfTest = fc.idf(tfTest,testing)
    print 'Counting TFxIDF...'
    tfIDFTest = fc.tfIDF(tfTest,idfTest,testing)
    label = fc.testing(tfIDFTest,W1,W2,B1,B2)
    hLoss = fc.hammingLoss(label,testing)
    hammingLoss.append(hLoss)
    print 'Hamming Loss = ', hLoss
    print
    print '================================================'
    print

print 'Mean hLoss = ' , np.mean(hammingLoss)
print time.time() - start_time , ' seconds'