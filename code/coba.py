from openpyxl import load_workbook
import re
import numpy as np
import math

def getData(sheet):
	dataSet = []
	for i in range(1, sheet.max_row+1):
		dataText = []
		label = []
		data = []
		dataText.append(sheet.cell(row=i, column=1).value)
		label.append(int(sheet.cell(row=i, column=2).value))
		label.append(int(sheet.cell(row=i, column=3).value))
		label.append(int(sheet.cell(row=i, column=4).value))
		data.append(dataText)
		data.append(label)
		dataSet.append(data)

	return dataSet

def preprocessing(dataSet):
	data = []
	for i in dataSet:
		data.append(i[0])
	f = open('stopwords.txt','r')
	stopwords = f.read()

	def stop_words(data):
		textStop = []
		print 'Now Start Remove Stopwords...'
		for i in data:
			for x in i[0].lower().split():
				x = re.sub('[(;:,.\'`?!0123456789)]', '', x)
				if x not in stopwords:
					textStop.append(x)
		return textStop

	dataStop  = stop_words(data)
	dataStop = np.unique(dataStop)
	thefile = open('dataStop.txt','w')
	for item in dataStop:
		print>>thefile, item
	print 'Preprocessing file saved'
	return dataStop	

def docEachClass(matriksTarget,dataSet,dataPre):
	data = []
	for i in matriksTarget:
		d = []
		count = 0
		for x in dataSet:
			if i == x[1]:
				count = count + 1
		d.append(i)
		d.append(count) #smoothing ditambahkan dengan banyaknya kata/fitur
		data.append(d)
	print 'Finish'
	return data

def docGivenWord(matriksTarget,dataPre,dataSet):
	result = []
	for i in dataPre:
		for k in matriksTarget:
			data = []
			countDoc = 0
			for j in dataSet:
				if k == j[1]:
					c = 0
					words = []
					for word in j[0][0].lower().split():
						word = re.sub('[(;:,.\'`?!0123456789)]', '', word)
						words.append(word)
					c = list(words).count(i)
					if c > 0:
						countDoc = countDoc + 1
			data.append(i)
			data.append(k)
			data.append(countDoc+1) #smoothing ditambahkan dengan 1
			result.append(data)
	thefile = open('Fw.txt','w')
	for item in result:
		print>>thefile, item
	print 'Finish'
	return result


def probEachClassGivenWord(docEachClass,Fw):
	result = []
	for i in Fw:
		for j in docEachClass:
			a = []
			if i[1] == j[0]:
				piW = (i[2]) / float(j[1])
			a.append(i[0])
			a.append(i[1])
			a.append(piW)
		result.append(a)
	print 'Finish'
	return result

def informationGain(piW,docEachClass,Fw,dataPre,dataSet):
	b = 0
	totalDoc = len(dataPre)*len(docEachClass)
	result = []
	for i in piW:
		for j in docEachClass:
			a = []
			if i[1] == j[0]:
				first = -((j[1])/float(totalDoc+len(dataSet))) * math.log((j[1])/float(totalDoc+len(dataSet)))
				second = Fw[b][2] * (i[2] * math.log(i[2]))
				third = (1 - Fw[b][2]) * ((1 - i[2]) * math.log(1 - i[2]))
				nilai = first + second + third
			a.append(i[0])
			a.append(i[1])
			a.append(nilai)
		result.append(a)
		b = b + 1
	print 'Finish'
	return result

matriksTarget = [[0,0,1],[0,1,0],[0,1,1],[1,0,0],[1,0,1],[1,1,0],[1,1,1]]
print 'Open File...'
wb = load_workbook('data.xlsx')
sheet = wb.active
dataSet = getData(sheet)
dataPre = preprocessing(dataSet)
print 'Count document for each class...'
docEachClass = docEachClass(matriksTarget,dataSet,dataPre)
print 'Count document for each class given word...'
Fw = docGivenWord(matriksTarget,dataPre,dataSet)
print 'Count Probability for Each Class Given Word w...'
piW = probEachClassGivenWord(docEachClass,Fw)
# print 'Count Information Gain for Each Class and Word...'
# igW = informationGain(piW,docEachClass,Fw,dataPre,dataSet)

for i in docEachClass:
	print i

# for i in Fw:
# 	print i

for i in piW:
	print i

# for i in igW:
# 	print i