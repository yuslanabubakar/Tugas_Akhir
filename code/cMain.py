# -*- coding: utf-8 -*-
"""
Created on Wed May 09 11:40:38 2018

@author: M. Yuslan Abubakar
"""

from openpyxl import load_workbook
import cFunction as fc
import time
import numpy as np

start_time = time.time()
label = ['anjuran','larangan','informasi']
print 'Open File...'
wb = load_workbook('dataMulti.xlsx')
sheet = wb.active
dataSet = fc.getData(sheet)

hammingLoss = []
num_folds = 5
subset_size = len(dataSet) / num_folds
for i in range(num_folds): #K-Cross Validation
    print 'K-Cross = ',(i+1)
    testing = dataSet[i*subset_size:][:subset_size]
    training = dataSet[:i*subset_size]+dataSet[(i+1)*subset_size:]
    print 'Now Start Remove Stopwords...'
    preprocessing = fc.preprocessing(training)
    print 'Counting document for each class...'
    docEachClass = fc.docEachClass(label,training)
    print 'Counting document for each class given word...'
    FW = fc.docGivenWord(training,preprocessing,label)
    print 'Counting Probability for Each Class Given Word w...'
    PIW = fc.probEachClassGivenWord(FW, label, docEachClass)
    print 'Counting Information Gain for Each Class and Word...'
    IGW = fc.informationGain(preprocessing,PIW,label,docEachClass,training)
    
    ##plot feature for each label
    #fc.plotIGW(IGW[0]) #anjuran
    #fc.plotIGW(IGW[1]) #larangan
    #fc.plotIGW(IGW[2]) #informasi
    
    #get threshold for each IGW
    value = 0.85
    igw = fc.thresholdIGW(IGW,value)
    
    ##for i in ig:
    ##    igWThreshold.append(i[0])
    ##np.savetxt("igWThreshold.txt", igWThreshold, delimiter=",", fmt="%s")
    print 'Counting TF...'
    TF = fc.tf(igw,training)
    print 'Counting IDF...'
    IDF = fc.idf(TF,training)
    print 'Counting TFxIDF...'
    TFIDF = fc.tfIDF(TF,training,IDF)
    
    print '================================================'
    
    #start training ann (artificial neural network)
    input_p = len(TFIDF[0][1])
    hidden_p = 10
    output_p = 1
    lr = 0.1
    epoch = 1000
    mseStandar = 0.01
    ##i_param = [input_p,hidden_p,output_p,lr,epoch,mseStandar]
    ##np.savetxt('inputParameters.txt',i_param)
    print 'Start training data...'
    W1Anjur,W2Anjur,B1Anjur,B2Anjur = fc.trainingAnjur(input_p,hidden_p,output_p,lr,epoch,mseStandar,TFIDF,training)
    W1Larang,W2Larang,B1Larang,B2Larang = fc.trainingLarang(input_p,hidden_p,output_p,lr,epoch,mseStandar,TFIDF,training)
    W1Info,W2Info,B1Info,B2Info = fc.trainingInfo(input_p,hidden_p,output_p,lr,epoch,mseStandar,TFIDF,training)
    print 'Finish'
    
    print '================================================'
    
    #start testing
    print 'Now is testing the data...'
#    wbTest = load_workbook('dataTest.xlsx')
#    sheetTest = wbTest.active
#    dataSetTest = fc.getData(sheetTest)
    print 'Counting TF...'
    tfTest = fc.tf(igw,testing)
    print 'Counting IDF...'
    idfTest = fc.idf(tfTest,testing)
    print 'Counting TFxIDF...'
    tfIDFTest = fc.tfIDF(tfTest,testing,idfTest)
    labelAnjur = fc.testing(tfIDFTest,W1Anjur,W2Anjur,B1Anjur,B2Anjur)
    labelLarang = fc.testing(tfIDFTest,W1Larang,W2Larang,B1Larang,B2Larang)
    labelInfo = fc.testing(tfIDFTest,W1Info,W2Info,B1Info,B2Info)
    print 'Count hamming loss...'
    hLoss = fc.hammingLoss(labelAnjur,labelLarang,labelInfo,testing,label)
    hammingLoss.append(hLoss)
    print 'Hamming Loss = ', hLoss
    print
    print '================================================'
    print

print 'Mean hLoss = ' , np.mean(hammingLoss)
print time.time() - start_time , ' seconds'