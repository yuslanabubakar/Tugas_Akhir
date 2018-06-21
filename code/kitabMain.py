# -*- coding: utf-8 -*-
"""
Created on Mon May 21 11:19:18 2018

@author: M. Yuslan Abubakar
"""

from openpyxl import load_workbook
import kitabFunction as fc
import time
import numpy as np

start_time = time.time()
matriksTarget = [[1,0,0,0,0],
                 [0,1,0,0,0],
                 [0,0,1,0,0],
                 [0,0,0,1,0],
                 [0,0,0,0,1]]
#np.savetxt('matriksTarget.txt',matriksTarget)
print 'Open File...'
wb = load_workbook('dataKitab.xlsx')
sheet = wb.active
dataSet = fc.getData(sheet)

accuracy = []
num_folds = 5
subset_size = len(dataSet) / num_folds
for i in range(num_folds): #K-Cross Validation
    print 'K-Cross = ',(i+1)
    testing = dataSet[i*subset_size:][:subset_size]
    training = dataSet[:i*subset_size]+dataSet[(i+1)*subset_size:]
    print 'Now Start Remove Stopwords...'
    preprocessing = fc.preprocessing(training)
    print 'Counting document for each class...'
    docEachClass = fc.docEachClass(matriksTarget,training)
    print 'Counting document for each class given word...'
    FW = fc.docGivenWord(training,preprocessing,matriksTarget)
    print 'Counting Probability for Each Class Given Word w...'
    PIW = fc.probEachClassGivenWord(FW,matriksTarget,docEachClass)
    print 'Counting Information Gain for Each Class and Word...'
    IGW = fc.informationGain(PIW,training,preprocessing,docEachClass)
    
    #plot feature for each label
    #fc.plotIGW(IGW)
    #
    #get threshold for each IGW
    value = 0.80
    igw = fc.thresholdIGW(IGW,value)
    #
    ###for i in ig:
    ###    igWThreshold.append(i[0])
    np.savetxt("igThresholdKitab.txt", igw, delimiter=",", fmt="%s")
    print 'Counting TF...'
    TF = fc.tf(igw,training)
    print 'Counting IDF...'
    IDF = fc.idf(TF,training)
    print 'Counting TFxIDF...'
    TFIDF = fc.tfIDF(TF,training,IDF)
    #
    #print '================================================'
    #
    #start training ann (artificial neural network)
    input_p = len(TFIDF[0][1])
    hidden_p = 10
    output_p = 5
    lr = 0.1
    epoch = 3000
    mseStandar = 0.01
    ###i_param = [input_p,hidden_p,output_p,lr,epoch,mseStandar]
    ###np.savetxt('inputParameters.txt',i_param)
    print 'Start training data...'
    W1,W2,B1,B2 = fc.training(input_p,hidden_p,output_p,lr,epoch,mseStandar,TFIDF,training)
    print 'Finish'
    
    print '================================================'
    
    #start testing
    print 'Now is testing the data...'
#    wbTest = load_workbook('dataTestKitab.xlsx')
#    sheetTest = wbTest.active
#    dataSetTest = fc.getData(sheetTest)
    print 'Counting TF...'
    tfTest = fc.tf(igw,testing)
    print 'Counting IDF...'
    idfTest = fc.idf(tfTest,testing)
    print 'Counting TFxIDF...'
    tfIDFTest = fc.tfIDF(tfTest,testing,idfTest)
    label = fc.testing(tfIDFTest,W1,W2,B1,B2)
    print 'Count accuracy...'
#    akurasi = fc.akurasi(label,testing) #akurasi biasa
    akurasi = fc.f1_score(matriksTarget,label,testing) #f1-score
    accuracy.append(akurasi)
    print 'Akurasi = ' , akurasi , '%'
    print
    print '================================================'
    print

print 'Mean Accuracy = ' , np.mean(accuracy)
print time.time() - start_time , ' seconds'